import csv
import calendar
import io
import json
import os
import sqlite3
from datetime import datetime
from functools import wraps

from flask import Blueprint, abort, flash, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from werkzeug.security import check_password_hash, generate_password_hash


mu = Blueprint("mu", __name__, url_prefix="/mu")
_db_path = ""
_csrf_validator = None


def _connect():
    conn = sqlite3.connect(_db_path, timeout=20)
    conn.execute("PRAGMA busy_timeout=30000")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.row_factory = sqlite3.Row
    return conn


def _now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _clean(value, limit=150):
    return " ".join(str(value or "").strip().split())[:limit]


def _monthly_stock_report(conn, month_value):
    try:
        report_date = datetime.strptime(month_value or "", "%Y-%m")
    except ValueError:
        report_date = datetime.now()
    month_key = report_date.strftime("%Y-%m")
    days_in_month = calendar.monthrange(report_date.year, report_date.month)[1]
    products = conn.execute(
        """SELECT p.id,p.code,p.name,p.unit,p.safety_stock,
                  COALESCE(SUM(l.qty),0) AS stock
           FROM mu_products p LEFT JOIN mu_lots l ON l.product_id=p.id
           GROUP BY p.id ORDER BY p.code"""
    ).fetchall()
    daily_rows = conn.execute(
        """SELECT product_id,CAST(substr(created_at,9,2) AS INTEGER) AS day,
                  SUM(qty) AS qty
           FROM mu_transactions
           WHERE status='Completed' AND substr(created_at,1,7)=?
           GROUP BY product_id,day""",
        (month_key,),
    ).fetchall()
    daily_by_product = {}
    for row in daily_rows:
        daily_by_product.setdefault(row["product_id"], {})[row["day"]] = int(row["qty"])
    report_rows = []
    for index, product in enumerate(products, start=1):
        daily = daily_by_product.get(product["id"], {})
        issued = sum(daily.values())
        stock = int(product["stock"])
        safety_stock = int(product["safety_stock"])
        report_rows.append(
            {
                "number": index,
                "code": product["code"],
                "name": product["name"],
                "unit": product["unit"],
                "total": stock + issued,
                "issued": issued,
                "stock": stock,
                "safety_stock": safety_stock,
                "order_qty": max(safety_stock - stock, 0),
                "daily": daily,
            }
        )
    return {
        "month": month_key,
        "label": report_date.strftime("%B %Y"),
        "days": list(range(1, days_in_month + 1)),
        "rows": report_rows,
    }


def _valid_date(value, required=False):
    value = str(value or "").strip()
    if not value:
        return None if not required else ""
    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        return ""


def _require_csrf():
    if not _csrf_validator or not _csrf_validator():
        abort(400, description="CSRF token ไม่ถูกต้อง")


def _admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        allowed_roles = {"admin_mu", "superadmin"}
        if not session.get("admin_logged_in") or session.get("admin_role") not in allowed_roles:
            return redirect(url_for("index", admin="1"))
        return view(*args, **kwargs)
    return wrapped


def _current_admin_id():
    return str(session.get("admin_username") or "").strip()


def _admin_response(success, message, tab="stock", status=200):
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({"success": success, "message": message}), status
    flash(message, "success" if success else "danger")
    return redirect(url_for("mu.admin_dashboard", tab=tab))


def _next_product_code(conn):
    rows = conn.execute(
        "SELECT code FROM mu_products WHERE upper(code) LIKE 'MU-%'"
    ).fetchall()
    highest = 0
    for row in rows:
        suffix = str(row["code"] or "").rsplit("-", 1)[-1]
        if suffix.isdigit():
            highest = max(highest, int(suffix))
    return f"MU-{highest + 1:03d}"


def _next_lot_number(conn, product_id, received_date=None):
    product = conn.execute(
        "SELECT code FROM mu_products WHERE id=?", (product_id,)
    ).fetchone()
    if not product:
        raise ValueError("ไม่พบสินค้า")
    date_part = (received_date or datetime.now().strftime("%Y-%m-%d")).replace("-", "")
    prefix = f"{product['code']}-{date_part}-"
    rows = conn.execute(
        "SELECT lot_number FROM mu_lots WHERE product_id=? AND lot_number LIKE ?",
        (product_id, f"{prefix}%"),
    ).fetchall()
    highest = 0
    for row in rows:
        suffix = str(row["lot_number"] or "").rsplit("-", 1)[-1]
        if suffix.isdigit():
            highest = max(highest, int(suffix))
    return f"{prefix}{highest + 1:03d}"


def ensure_mu_schema():
    conn = _connect()
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS mu_settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS mu_admins (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE COLLATE NOCASE,
            password_hash TEXT NOT NULL,
            display_name TEXT NOT NULL,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS mu_user_access (
            emp_id TEXT PRIMARY KEY,
            password_hash TEXT NOT NULL,
            updated_by TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(emp_id) REFERENCES users(emp_id)
        );
        CREATE TABLE IF NOT EXISTS mu_products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL UNIQUE COLLATE NOCASE,
            name TEXT NOT NULL,
            unit TEXT NOT NULL,
            safety_stock INTEGER NOT NULL DEFAULT 0 CHECK(safety_stock >= 0),
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS mu_lots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            lot_number TEXT NOT NULL,
            qty INTEGER NOT NULL DEFAULT 0 CHECK(qty >= 0),
            received_date TEXT NOT NULL,
            expiry_date TEXT,
            created_by TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(product_id) REFERENCES mu_products(id)
        );
        CREATE TABLE IF NOT EXISTS mu_transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            emp_id TEXT NOT NULL,
            requester_name TEXT NOT NULL,
            qty INTEGER NOT NULL CHECK(qty > 0),
            status TEXT NOT NULL DEFAULT 'Completed',
            note TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            edited_by TEXT,
            cancelled_by TEXT,
            FOREIGN KEY(product_id) REFERENCES mu_products(id)
        );
        CREATE TABLE IF NOT EXISTS mu_transaction_lots (
            transaction_id INTEGER NOT NULL,
            lot_id INTEGER NOT NULL,
            qty INTEGER NOT NULL CHECK(qty > 0),
            PRIMARY KEY(transaction_id, lot_id),
            FOREIGN KEY(transaction_id) REFERENCES mu_transactions(id),
            FOREIGN KEY(lot_id) REFERENCES mu_lots(id)
        );
        CREATE TABLE IF NOT EXISTS mu_audit_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            actor_type TEXT NOT NULL,
            actor_id TEXT NOT NULL,
            action TEXT NOT NULL,
            entity_type TEXT NOT NULL,
            entity_id INTEGER,
            detail TEXT NOT NULL,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS mu_stock_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            lot_id INTEGER,
            event_type TEXT NOT NULL,
            qty INTEGER NOT NULL CHECK(qty > 0),
            stock_before INTEGER NOT NULL,
            stock_after INTEGER NOT NULL,
            reason TEXT NOT NULL,
            actor_id TEXT NOT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY(product_id) REFERENCES mu_products(id),
            FOREIGN KEY(lot_id) REFERENCES mu_lots(id)
        );
        CREATE INDEX IF NOT EXISTS idx_mu_lots_fifo
            ON mu_lots(product_id, received_date, id);
        CREATE INDEX IF NOT EXISTS idx_mu_transactions_history
            ON mu_transactions(created_at DESC, id DESC);
        CREATE INDEX IF NOT EXISTS idx_mu_audit_created
            ON mu_audit_logs(created_at DESC, id DESC);
        CREATE INDEX IF NOT EXISTS idx_mu_stock_events_created
            ON mu_stock_events(created_at DESC, id DESC);
        """
    )
    stamp = _now()
    seed_path = os.path.join(os.path.dirname(__file__), "mu_seed_items.csv")
    if os.path.exists(seed_path):
        with open(seed_path, newline="", encoding="utf-8-sig") as seed_file:
            for item in csv.DictReader(seed_file):
                conn.execute(
                    """INSERT OR IGNORE INTO mu_products
                       (code, name, unit, safety_stock, is_active, created_at, updated_at)
                       VALUES (?, ?, ?, ?, 1, ?, ?)""",
                    (item["code"], item["name"], item["unit"], int(item["safety_stock"]), stamp, stamp),
                )

    conn.commit()
    conn.close()


def register_mu_module(app, db_path, csrf_validator):
    global _db_path, _csrf_validator
    _db_path = db_path
    _csrf_validator = csrf_validator
    ensure_mu_schema()
    app.register_blueprint(mu)


def _user_context(emp_id):
    if session.get("user_id") != emp_id:
        return None
    conn = _connect()
    user = conn.execute(
        "SELECT emp_id,name,department,location FROM users WHERE emp_id=?", (emp_id,)
    ).fetchone()
    conn.close()
    if not user or str(user["department"] or "").strip().casefold() != "manufacturing":
        return None
    return user


def _available_stock(conn, product_id):
    return int(conn.execute(
        "SELECT COALESCE(SUM(qty),0) FROM mu_lots WHERE product_id=?", (product_id,)
    ).fetchone()[0])


def _consume_fifo(conn, product_id, qty):
    lots = conn.execute(
        """SELECT id,qty FROM mu_lots
           WHERE product_id=? AND qty>0
           ORDER BY date(received_date) ASC, id ASC""",
        (product_id,),
    ).fetchall()
    remaining = qty
    allocations = []
    for lot in lots:
        take = min(remaining, int(lot["qty"]))
        if take:
            conn.execute("UPDATE mu_lots SET qty=qty-?,updated_at=? WHERE id=?", (take, _now(), lot["id"]))
            allocations.append((lot["id"], take))
            remaining -= take
        if remaining == 0:
            break
    if remaining:
        raise ValueError("สต็อกไม่เพียงพอ")
    return allocations


def _restore_allocations(conn, transaction_id):
    rows = conn.execute(
        "SELECT lot_id,qty FROM mu_transaction_lots WHERE transaction_id=?", (transaction_id,)
    ).fetchall()
    for row in rows:
        conn.execute("UPDATE mu_lots SET qty=qty+?,updated_at=? WHERE id=?", (row["qty"], _now(), row["lot_id"]))
    conn.execute("DELETE FROM mu_transaction_lots WHERE transaction_id=?", (transaction_id,))


def _save_allocations(conn, transaction_id, allocations):
    conn.executemany(
        "INSERT INTO mu_transaction_lots(transaction_id,lot_id,qty) VALUES(?,?,?)",
        [(transaction_id, lot_id, qty) for lot_id, qty in allocations],
    )


@mu.route("/access/<emp_id>", methods=["GET", "POST"])
def access(emp_id):
    user = _user_context(emp_id)
    if not user:
        flash("กรุณาเข้าสู่ระบบพนักงานใหม่", "danger")
        return redirect(url_for("index"))
    conn = _connect()
    setting = conn.execute(
        "SELECT password_hash FROM mu_user_access WHERE emp_id=?", (emp_id,)
    ).fetchone()
    conn.close()
    if request.method == "POST":
        _require_csrf()
        password = request.form.get("password", "")
        if not setting:
            flash("ยังไม่มีสิทธิ์ใช้งาน MU เนื่องจาก Admin MU ยังไม่ได้ตั้งรหัสผ่านให้บัญชีนี้", "warning")
        elif check_password_hash(setting["password_hash"], password):
            session["mu_user_emp_id"] = emp_id
            if request.form.get("next") == "admin":
                return redirect(url_for("mu.admin_dashboard"))
            return redirect(url_for("mu.portal", emp_id=emp_id))
        else:
            flash("รหัสผ่าน MU ไม่ถูกต้อง", "danger")
    return render_template(
        "mu_access.html",
        user=user,
        has_mu_access=bool(setting),
        next_page="admin" if request.args.get("next") == "admin" else "",
    )


@mu.route("/portal/<emp_id>")
def portal(emp_id):
    user = _user_context(emp_id)
    if not user or session.get("mu_user_emp_id") != emp_id:
        return redirect(url_for("mu.access", emp_id=emp_id))
    search = _clean(request.args.get("q"), 80)
    conn = _connect()
    params = []
    where = "WHERE p.is_active=1"
    if search:
        where += " AND (p.code LIKE ? OR p.name LIKE ?)"
        params.extend([f"%{search}%", f"%{search}%"])
    products = conn.execute(
        f"""SELECT p.*,COALESCE(SUM(l.qty),0) AS stock
            FROM mu_products p LEFT JOIN mu_lots l ON l.product_id=p.id
            {where} GROUP BY p.id ORDER BY p.code""",
        params,
    ).fetchall()
    history = conn.execute(
        """SELECT t.*,p.code,p.name,p.unit FROM mu_transactions t
           JOIN mu_products p ON p.id=t.product_id
           WHERE t.emp_id=? ORDER BY t.created_at DESC,t.id DESC LIMIT 20""",
        (emp_id,),
    ).fetchall()
    conn.close()
    return render_template("mu_portal.html", user=user, products=products, history=history, search=search)


@mu.route("/withdraw/<emp_id>", methods=["POST"])
def withdraw(emp_id):
    _require_csrf()
    user = _user_context(emp_id)
    if not user or session.get("mu_user_emp_id") != emp_id:
        abort(403)
    product_id = request.form.get("product_id", type=int)
    qty = request.form.get("qty", type=int)
    note = _clean(request.form.get("note"), 300)
    if not product_id or not qty or qty <= 0 or qty > 100000:
        flash("จำนวนเบิกไม่ถูกต้อง", "danger")
        return redirect(url_for("mu.portal", emp_id=emp_id))
    conn = _connect()
    try:
        conn.execute("BEGIN IMMEDIATE")
        product = conn.execute("SELECT id,name FROM mu_products WHERE id=? AND is_active=1", (product_id,)).fetchone()
        if not product or _available_stock(conn, product_id) < qty:
            raise ValueError("สต็อกไม่เพียงพอ")
        stamp = _now()
        cursor = conn.execute(
            """INSERT INTO mu_transactions
               (product_id,emp_id,requester_name,qty,status,note,created_at,updated_at)
               VALUES(?,?,?,?, 'Completed',?,?,?)""",
            (product_id, emp_id, user["name"], qty, note, stamp, stamp),
        )
        transaction_id = cursor.lastrowid
        _save_allocations(conn, transaction_id, _consume_fifo(conn, product_id, qty))
        conn.execute(
            """INSERT INTO mu_audit_logs(actor_type,actor_id,action,entity_type,entity_id,detail,created_at)
               VALUES('user',?,'withdraw','transaction',?,?,?)""",
            (emp_id, transaction_id, f"เบิก {product['name']} จำนวน {qty}", stamp),
        )
        conn.commit()
        flash("เบิกสำเร็จและตัดสต็อกแบบ FIFO แล้ว", "success")
    except (ValueError, sqlite3.Error) as exc:
        conn.rollback()
        flash(str(exc), "danger")
    finally:
        conn.close()
    return redirect(url_for("mu.portal", emp_id=emp_id))


@mu.route("/withdraw-batch/<emp_id>", methods=["POST"])
def withdraw_batch(emp_id):
    """เบิกหลายรายการในคำสั่งเดียวและ rollback ทั้งชุดหากรายการใดไม่ผ่าน"""
    _require_csrf()
    user = _user_context(emp_id)
    if not user or session.get("mu_user_emp_id") != emp_id:
        abort(403)
    try:
        raw_items = json.loads(request.form.get("items", "[]"))
    except (TypeError, ValueError, json.JSONDecodeError):
        raw_items = []
    requested = {}
    for item in raw_items if isinstance(raw_items, list) else []:
        try:
            product_id = int(item.get("productId", 0))
            qty = int(item.get("qty", 0))
        except (TypeError, ValueError, AttributeError):
            continue
        if product_id > 0 and 0 < qty <= 100000:
            requested[product_id] = requested.get(product_id, 0) + qty
    if not requested or len(requested) > 100:
        flash("ตะกร้าเบิกไม่ถูกต้องหรือไม่มีรายการ", "danger")
        return redirect(url_for("mu.portal", emp_id=emp_id))

    conn = _connect()
    try:
        conn.execute("BEGIN IMMEDIATE")
        stamp = _now()
        batch_reference = f"MU-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        for product_id, qty in requested.items():
            product = conn.execute(
                "SELECT id,code,name FROM mu_products WHERE id=? AND is_active=1",
                (product_id,),
            ).fetchone()
            if not product or _available_stock(conn, product_id) < qty:
                raise ValueError(
                    f"สต็อกไม่เพียงพอ: {product['name'] if product else product_id}"
                )
            cursor = conn.execute(
                """INSERT INTO mu_transactions
                   (product_id,emp_id,requester_name,qty,status,note,created_at,updated_at)
                   VALUES(?,?,?,?, 'Completed',?,?,?)""",
                (
                    product_id,
                    emp_id,
                    user["name"],
                    qty,
                    f"เบิกจากตะกร้า {batch_reference}",
                    stamp,
                    stamp,
                ),
            )
            transaction_id = cursor.lastrowid
            _save_allocations(
                conn, transaction_id, _consume_fifo(conn, product_id, qty)
            )
            conn.execute(
                """INSERT INTO mu_audit_logs
                   (actor_type,actor_id,action,entity_type,entity_id,detail,created_at)
                   VALUES('user',?,'withdraw','transaction',?,?,?)""",
                (
                    emp_id,
                    transaction_id,
                    f"{batch_reference}: เบิก {product['name']} จำนวน {qty}",
                    stamp,
                ),
            )
        conn.commit()
        flash(
            f"เบิกสำเร็จ {len(requested)} รายการ และตัดสต็อกแบบ FIFO แล้ว",
            "success",
        )
    except (ValueError, sqlite3.Error) as exc:
        conn.rollback()
        flash(str(exc), "danger")
    finally:
        conn.close()
    return redirect(url_for("mu.portal", emp_id=emp_id))


@mu.route("/admin/logout", methods=["POST"])
@_admin_required
def admin_logout():
    _require_csrf()
    session.clear()
    return redirect(url_for("index", admin="1"))


@mu.route("/admin")
@_admin_required
def admin_dashboard():
    tab = request.args.get("tab", "stock")
    if tab not in {"stock", "history", "audit", "monthly", "employees", "profile", "admins"}:
        tab = "stock"
    try:
        product_page = max(1, int(request.args.get("page", 1)))
    except (TypeError, ValueError):
        product_page = 1
    products_per_page = 25
    conn = _connect()
    product_count = int(conn.execute("SELECT COUNT(*) FROM mu_products").fetchone()[0])
    total_product_pages = max(1, (product_count + products_per_page - 1) // products_per_page)
    product_page = min(product_page, total_product_pages)
    products = conn.execute(
        """SELECT p.*,COALESCE(SUM(l.qty),0) AS stock,COUNT(l.id) AS lot_count
           FROM mu_products p LEFT JOIN mu_lots l ON l.product_id=p.id
           GROUP BY p.id ORDER BY p.code
           LIMIT ? OFFSET ?""",
        (products_per_page, (product_page - 1) * products_per_page),
    ).fetchall()
    total_stock = int(
        conn.execute("SELECT COALESCE(SUM(qty),0) FROM mu_lots").fetchone()[0]
    )
    out_of_stock_count = int(
        conn.execute(
            """SELECT COUNT(*) FROM mu_products p
               WHERE COALESCE((SELECT SUM(l.qty) FROM mu_lots l WHERE l.product_id=p.id),0)<=0"""
        ).fetchone()[0]
    )
    history = conn.execute(
        """SELECT t.*,p.code,p.name,p.unit FROM mu_transactions t
           JOIN mu_products p ON p.id=t.product_id
           ORDER BY t.created_at DESC,t.id DESC LIMIT 200"""
    ).fetchall()
    audit = conn.execute(
        """SELECT a.*,
                  t.created_at AS transaction_date,t.requester_name,t.emp_id,t.qty AS transaction_qty,
                  p.code AS product_code,p.name AS product_name,p.unit AS product_unit,
                  GROUP_CONCAT(l.lot_number || ' (' || tl.qty || ')', ', ') AS lot_detail
           FROM mu_audit_logs a
           LEFT JOIN mu_transactions t
             ON a.entity_type='transaction' AND t.id=a.entity_id
           LEFT JOIN mu_products p ON p.id=t.product_id
           LEFT JOIN mu_transaction_lots tl ON tl.transaction_id=t.id
           LEFT JOIN mu_lots l ON l.id=tl.lot_id
           GROUP BY a.id
           ORDER BY a.created_at DESC,a.id DESC LIMIT 200"""
    ).fetchall()
    stock_events = conn.execute(
        """SELECT e.*,p.code,p.name,p.unit,l.lot_number
           FROM mu_stock_events e
           JOIN mu_products p ON p.id=e.product_id
           LEFT JOIN mu_lots l ON l.id=e.lot_id
           ORDER BY e.created_at DESC,e.id DESC LIMIT 200"""
    ).fetchall()
    employees = conn.execute(
        """SELECT u.emp_id,u.name,u.department,u.location,
                  CASE WHEN a.emp_id IS NULL THEN 0 ELSE 1 END AS has_mu_password,
                  a.updated_by,a.updated_at
           FROM users u
           LEFT JOIN mu_user_access a ON a.emp_id=u.emp_id
           WHERE lower(trim(COALESCE(u.department,'')))='manufacturing'
           ORDER BY u.emp_id"""
    ).fetchall()
    admin_profile = conn.execute(
        "SELECT id,username,name,role FROM admins WHERE username=?",
        (_current_admin_id(),),
    ).fetchone()
    mu_admins = conn.execute(
        "SELECT id,username,name,role FROM admins WHERE role='admin_mu' ORDER BY username"
    ).fetchall()
    monthly_report = _monthly_stock_report(conn, request.args.get("month"))
    summary = {
        "products": product_count,
        "stock": total_stock,
        "out_of_stock": out_of_stock_count,
        "completed": sum(1 for row in history if row["status"] == "Completed"),
        "cancelled": sum(1 for row in history if row["status"] == "Cancelled"),
    }
    pagination = {
        "page": product_page,
        "per_page": products_per_page,
        "total": product_count,
        "pages": total_product_pages,
        "start": ((product_page - 1) * products_per_page + 1) if product_count else 0,
        "end": min(product_page * products_per_page, product_count),
    }
    conn.close()
    return render_template(
        "mu_admin.html", tab=tab, products=products, history=history,
        audit=audit, stock_events=stock_events, employees=employees, summary=summary,
        admin_profile=admin_profile, mu_admins=mu_admins, pagination=pagination,
        monthly_report=monthly_report,
    )


@mu.route("/admin/next-product-code")
@_admin_required
def next_product_code():
    conn = _connect()
    try:
        return jsonify({"success": True, "next_code": _next_product_code(conn)})
    finally:
        conn.close()


@mu.route("/admin/product/<int:product_id>/next-lot")
@_admin_required
def next_lot_number(product_id):
    received_date = _valid_date(request.args.get("received_date")) or None
    conn = _connect()
    try:
        return jsonify(
            {
                "success": True,
                "next_lot": _next_lot_number(conn, product_id, received_date),
            }
        )
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 404
    finally:
        conn.close()


@mu.route("/admin/audit/export.xlsx")
@_admin_required
def export_audit_excel():
    conn = _connect()
    try:
        withdrawals = conn.execute(
            """SELECT t.created_at,t.requester_name,t.emp_id,p.code,p.name,t.qty,p.unit,
                      t.status,COALESCE(t.note,'') AS note,
                      GROUP_CONCAT(l.lot_number || ' (' || tl.qty || ')', ', ') AS lots
               FROM mu_transactions t
               JOIN mu_products p ON p.id=t.product_id
               LEFT JOIN mu_transaction_lots tl ON tl.transaction_id=t.id
               LEFT JOIN mu_lots l ON l.id=tl.lot_id
               GROUP BY t.id ORDER BY t.created_at DESC,t.id DESC"""
        ).fetchall()
        stock_events = conn.execute(
            """SELECT e.created_at,p.code,p.name,COALESCE(l.lot_number,'-') AS lot_number,
                      e.event_type,e.qty,p.unit,e.stock_before,e.stock_after,e.reason,e.actor_id
               FROM mu_stock_events e
               JOIN mu_products p ON p.id=e.product_id
               LEFT JOIN mu_lots l ON l.id=e.lot_id
               ORDER BY e.created_at DESC,e.id DESC"""
        ).fetchall()
        audit_rows = conn.execute(
            """SELECT created_at,actor_type,actor_id,action,entity_type,entity_id,detail
               FROM mu_audit_logs ORDER BY created_at DESC,id DESC"""
        ).fetchall()
        monthly_report = _monthly_stock_report(conn, request.args.get("month"))
    finally:
        conn.close()

    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="321FDB")
    header_font = Font(color="FFFFFF", bold=True)

    def safe_excel_value(value):
        if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
            return f"'{value}"
        return value

    def add_sheet(title, headers, rows):
        sheet = workbook.create_sheet(title)
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            sheet.append([safe_excel_value(value) for value in row])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            width = min(45, max(len(str(cell.value or "")) for cell in column) + 2)
            sheet.column_dimensions[column[0].column_letter].width = width
        return sheet

    add_sheet(
        "รายการเบิก",
        ["วันเวลา", "ผู้เบิก", "รหัสพนักงาน", "รหัสสินค้า", "สินค้า", "จำนวน", "หน่วย", "สถานะ", "หมายเหตุ", "Lot FIFO"],
        [tuple(row) for row in withdrawals],
    )
    add_sheet(
        "Stock Movement",
        ["วันเวลา", "รหัสสินค้า", "สินค้า", "Lot", "ประเภท", "จำนวน", "หน่วย", "ก่อน", "หลัง", "เหตุผล", "ผู้ดำเนินการ"],
        [tuple(row) for row in stock_events],
    )
    add_sheet(
        "Audit Log",
        ["วันเวลา", "ประเภทผู้ใช้", "ผู้ดำเนินการ", "เหตุการณ์", "ประเภทข้อมูล", "รหัสข้อมูล", "รายละเอียด"],
        [tuple(row) for row in audit_rows],
    )
    monthly_headers = [
        "No.", "รหัสสินค้า", "รายการ / List", "รวม", "ยอดเบิก",
        "คงเหลือ", "Safety Stock", "สั่งซื้อ", "หน่วย",
    ] + [str(day) for day in monthly_report["days"]]
    monthly_rows = [
        [
            row["number"], row["code"], row["name"], row["total"], row["issued"],
            row["stock"], row["safety_stock"], row["order_qty"], row["unit"],
        ] + [row["daily"].get(day, "") for day in monthly_report["days"]]
        for row in monthly_report["rows"]
    ]
    monthly_sheet = add_sheet(
        f"รายเดือน {monthly_report['month']}",
        monthly_headers,
        monthly_rows,
    )
    monthly_sheet.freeze_panes = "J2"
    out_fill = PatternFill("solid", fgColor="F8D7DA")
    low_fill = PatternFill("solid", fgColor="FFF3CD")
    thin_side = Side(style="thin", color="B7C1CE")
    section_side = Side(style="medium", color="667085")
    monthly_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    monthly_sheet.sheet_view.showGridLines = False
    monthly_sheet.row_dimensions[1].height = 30
    monthly_widths = {
        "A": 7, "B": 15, "C": 36, "D": 10, "E": 10,
        "F": 11, "G": 13, "H": 10, "I": 11,
    }
    for column_letter, width in monthly_widths.items():
        monthly_sheet.column_dimensions[column_letter].width = width
    for column_index in range(10, 10 + len(monthly_report["days"])):
        monthly_sheet.column_dimensions[
            monthly_sheet.cell(1, column_index).column_letter
        ].width = 4.5
    for cell in monthly_sheet[1]:
        cell.fill = PatternFill("solid", fgColor="263D8F")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = monthly_border
    monthly_sheet["I1"].border = Border(
        left=thin_side, right=section_side, top=thin_side, bottom=thin_side
    )
    for excel_row, report_row in enumerate(monthly_report["rows"], start=2):
        fill = out_fill if report_row["stock"] <= 0 else (
            low_fill if report_row["stock"] <= report_row["safety_stock"] else None
        )
        monthly_sheet.row_dimensions[excel_row].height = 22
        for cell in monthly_sheet[excel_row]:
            cell.border = monthly_border
            cell.alignment = Alignment(
                horizontal="left" if cell.column == 3 else "center",
                vertical="center",
                wrap_text=cell.column == 3,
            )
            if fill:
                cell.fill = fill
        monthly_sheet.cell(excel_row, 9).border = Border(
            left=thin_side, right=section_side, top=thin_side, bottom=thin_side
        )
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"MU_Stock_Audit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@mu.route("/admin/product", methods=["POST"])
@_admin_required
def add_product():
    _require_csrf()
    code = _clean(request.form.get("code"), 30).upper()
    name = _clean(request.form.get("name"), 150)
    unit = _clean(request.form.get("unit"), 30)
    safety = request.form.get("safety_stock", type=int)
    initial_qty = request.form.get("initial_qty", default=0, type=int)
    received_date = _valid_date(
        request.form.get("received_date"), required=bool(initial_qty and initial_qty > 0)
    )
    expiry_date = _valid_date(request.form.get("expiry_date")) or None
    if (
        not name or not unit or safety is None or safety < 0
        or initial_qty is None or initial_qty < 0
        or (initial_qty > 0 and not received_date)
        or (expiry_date and received_date and expiry_date < received_date)
    ):
        return _admin_response(False, "ข้อมูลสินค้าไม่ครบหรือไม่ถูกต้อง", status=400)
    conn = _connect()
    try:
        conn.execute("BEGIN IMMEDIATE")
        stamp = _now()
        if not code:
            code = _next_product_code(conn)
        cursor = conn.execute(
            """INSERT INTO mu_products(code,name,unit,safety_stock,is_active,created_at,updated_at)
               VALUES(?,?,?,?,1,?,?)""", (code, name, unit, safety, stamp, stamp)
        )
        conn.execute(
            """INSERT INTO mu_audit_logs(actor_type,actor_id,action,entity_type,entity_id,detail,created_at)
               VALUES('admin',?,'create','product',?,?,?)""",
            (_current_admin_id(), cursor.lastrowid, f"เพิ่มสินค้า {code} {name}", stamp),
        )
        lot_number = None
        if initial_qty > 0:
            lot_number = _next_lot_number(conn, cursor.lastrowid, received_date)
            lot_cursor = conn.execute(
                """INSERT INTO mu_lots
                   (product_id,lot_number,qty,received_date,expiry_date,created_by,created_at,updated_at)
                   VALUES(?,?,?,?,?,?,?,?)""",
                (
                    cursor.lastrowid, lot_number, initial_qty, received_date,
                    expiry_date, _current_admin_id(), stamp, stamp,
                ),
            )
            conn.execute(
                """INSERT INTO mu_audit_logs
                   (actor_type,actor_id,action,entity_type,entity_id,detail,created_at)
                   VALUES('admin',?,'receive','lot',?,?,?)""",
                (
                    _current_admin_id(), lot_cursor.lastrowid,
                    f"รับ Lot ตั้งต้น {lot_number} จำนวน {initial_qty}", stamp,
                ),
            )
        conn.commit()
        message = f"เพิ่มสินค้า {code} แล้ว"
        if lot_number:
            message += f" พร้อม Lot ตั้งต้น {lot_number} จำนวน {initial_qty} {unit}"
        return _admin_response(True, message)
    except sqlite3.IntegrityError:
        conn.rollback()
        return _admin_response(False, "รหัสสินค้านี้มีอยู่แล้ว", status=409)
    except (ValueError, sqlite3.Error) as exc:
        conn.rollback()
        return _admin_response(False, str(exc) or "ไม่สามารถเพิ่มสินค้าได้", status=400)
    finally:
        conn.close()


@mu.route("/admin/product/<int:product_id>/edit", methods=["POST"])
@_admin_required
def edit_product(product_id):
    _require_csrf()
    code = _clean(request.form.get("code"), 30).upper()
    name = _clean(request.form.get("name"), 150)
    unit = _clean(request.form.get("unit"), 30)
    safety = request.form.get("safety_stock", type=int)
    if not code or not name or not unit or safety is None or safety < 0:
        return _admin_response(False, "ข้อมูลสินค้าไม่ครบหรือไม่ถูกต้อง", status=400)
    conn = _connect()
    try:
        product = conn.execute("SELECT * FROM mu_products WHERE id=?", (product_id,)).fetchone()
        if not product:
            raise ValueError("ไม่พบสินค้าที่ต้องการแก้ไข")
        stamp = _now()
        conn.execute(
            """UPDATE mu_products SET code=?,name=?,unit=?,safety_stock=?,updated_at=?
               WHERE id=?""",
            (code, name, unit, safety, stamp, product_id),
        )
        conn.execute(
            """INSERT INTO mu_audit_logs
               (actor_type,actor_id,action,entity_type,entity_id,detail,created_at)
               VALUES('admin',?,'edit_product','product',?,?,?)""",
            (
                _current_admin_id(),
                product_id,
                f"แก้สินค้า {product['code']} -> {code}; ชื่อ {product['name']} -> {name}; Safety {product['safety_stock']} -> {safety}",
                stamp,
            ),
        )
        conn.commit()
        return _admin_response(True, "แก้ไขข้อมูลสินค้าแล้ว")
    except sqlite3.IntegrityError:
        conn.rollback()
        return _admin_response(False, "รหัสสินค้านี้ถูกใช้แล้ว", status=409)
    except ValueError as exc:
        conn.rollback()
        return _admin_response(False, str(exc), status=404)
    finally:
        conn.close()


@mu.route("/admin/product/<int:product_id>/toggle", methods=["POST"])
@_admin_required
def toggle_product(product_id):
    _require_csrf()
    conn = _connect()
    try:
        product = conn.execute("SELECT id,code,name,is_active FROM mu_products WHERE id=?", (product_id,)).fetchone()
        if not product:
            raise ValueError("ไม่พบสินค้า")
        new_status = 0 if product["is_active"] else 1
        stamp = _now()
        conn.execute("UPDATE mu_products SET is_active=?,updated_at=? WHERE id=?", (new_status, stamp, product_id))
        conn.execute(
            """INSERT INTO mu_audit_logs
               (actor_type,actor_id,action,entity_type,entity_id,detail,created_at)
               VALUES('admin',?,'toggle_product','product',?,?,?)""",
            (_current_admin_id(), product_id, f"{'เปิด' if new_status else 'ปิด'}ใช้งาน {product['code']} {product['name']}", stamp),
        )
        conn.commit()
        return _admin_response(True, f"{'เปิด' if new_status else 'ปิด'}ใช้งานสินค้าแล้ว")
    except ValueError as exc:
        conn.rollback()
        return _admin_response(False, str(exc), status=404)
    finally:
        conn.close()


@mu.route("/admin/lot", methods=["POST"])
@_admin_required
def add_lot():
    _require_csrf()
    product_id = request.form.get("product_id", type=int)
    qty = request.form.get("qty", type=int)
    lot_number = _clean(request.form.get("lot_number"), 50)
    received_date = _valid_date(request.form.get("received_date"), required=True)
    expiry_date = _valid_date(request.form.get("expiry_date")) or None
    if not product_id or not qty or qty <= 0 or not received_date:
        return _admin_response(False, "ข้อมูล Lot ไม่ครบหรือไม่ถูกต้อง", status=400)
    conn = _connect()
    try:
        conn.execute("BEGIN IMMEDIATE")
        stamp = _now()
        if not lot_number:
            lot_number = _next_lot_number(conn, product_id, received_date)
        duplicate = conn.execute(
            "SELECT id FROM mu_lots WHERE product_id=? AND upper(lot_number)=upper(?)",
            (product_id, lot_number),
        ).fetchone()
        if duplicate:
            # กรณีผู้ใช้เปิด popup ค้างไว้ ระบบจะขยับ sequence ให้ใหม่ก่อนบันทึก
            lot_number = _next_lot_number(conn, product_id, received_date)
        cursor = conn.execute(
            """INSERT INTO mu_lots(product_id,lot_number,qty,received_date,expiry_date,created_by,created_at,updated_at)
               VALUES(?,?,?,?,?,?,?,?)""",
            (product_id, lot_number, qty, received_date, expiry_date, _current_admin_id(), stamp, stamp),
        )
        conn.execute(
            """INSERT INTO mu_audit_logs(actor_type,actor_id,action,entity_type,entity_id,detail,created_at)
               VALUES('admin',?,'receive','lot',?,?,?)""",
            (_current_admin_id(), cursor.lastrowid, f"รับ Lot {lot_number} จำนวน {qty}", stamp),
        )
        conn.commit()
        return _admin_response(True, f"รับสินค้าเข้า Lot {lot_number} แล้ว")
    except (ValueError, sqlite3.Error) as exc:
        conn.rollback()
        return _admin_response(False, str(exc) or "ไม่สามารถเพิ่ม Lot ได้", status=400)
    finally:
        conn.close()


@mu.route("/admin/lot/<int:lot_id>/edit", methods=["POST"])
@_admin_required
def edit_lot(lot_id):
    _require_csrf()
    lot_number = _clean(request.form.get("lot_number"), 50)
    qty = request.form.get("qty", type=int)
    received_date = _valid_date(request.form.get("received_date"), required=True)
    expiry_date = _valid_date(request.form.get("expiry_date")) or None
    reason = _clean(request.form.get("reason"), 300)
    if not lot_number or qty is None or qty < 0 or not received_date or not reason:
        return _admin_response(False, "ข้อมูลแก้ไข Lot ไม่ครบหรือไม่ถูกต้อง", status=400)
    conn = _connect()
    try:
        conn.execute("BEGIN IMMEDIATE")
        lot = conn.execute(
            """SELECT l.*,p.code,p.name FROM mu_lots l
               JOIN mu_products p ON p.id=l.product_id WHERE l.id=?""",
            (lot_id,),
        ).fetchone()
        if not lot:
            raise ValueError("ไม่พบ Lot")
        allocated = conn.execute(
            "SELECT COALESCE(SUM(qty),0) FROM mu_transaction_lots WHERE lot_id=?",
            (lot_id,),
        ).fetchone()[0]
        if qty < 0 or (qty == 0 and allocated and not reason):
            raise ValueError("จำนวน Lot ไม่ถูกต้อง")
        before = _available_stock(conn, lot["product_id"])
        stamp = _now()
        conn.execute(
            """UPDATE mu_lots SET lot_number=?,qty=?,received_date=?,expiry_date=?,updated_at=?
               WHERE id=?""",
            (lot_number, qty, received_date, expiry_date, stamp, lot_id),
        )
        after = before - int(lot["qty"]) + qty
        difference = abs(qty - int(lot["qty"]))
        if difference:
            conn.execute(
                """INSERT INTO mu_stock_events
                   (product_id,lot_id,event_type,qty,stock_before,stock_after,reason,actor_id,created_at)
                   VALUES(?,?,'lot_adjust',?,?,?,?,?,?)""",
                (lot["product_id"], lot_id, difference, before, after, reason, _current_admin_id(), stamp),
            )
        conn.execute(
            """INSERT INTO mu_audit_logs
               (actor_type,actor_id,action,entity_type,entity_id,detail,created_at)
               VALUES('admin',?,'edit_lot','lot',?,?,?)""",
            (
                _current_admin_id(),
                lot_id,
                f"{lot['code']} Lot {lot['lot_number']} -> {lot_number}; จำนวน {lot['qty']} -> {qty}; เหตุผล: {reason}",
                stamp,
            ),
        )
        conn.commit()
        return _admin_response(True, "แก้ไข Lot แล้ว")
    except (ValueError, sqlite3.Error) as exc:
        conn.rollback()
        return _admin_response(False, str(exc), status=400)
    finally:
        conn.close()


@mu.route("/admin/lot/<int:lot_id>/delete", methods=["POST"])
@_admin_required
def delete_lot(lot_id):
    _require_csrf()
    reason = _clean(request.form.get("reason"), 300)
    if not reason:
        return _admin_response(False, "กรุณาระบุเหตุผลการลบ Lot", status=400)
    conn = _connect()
    try:
        conn.execute("BEGIN IMMEDIATE")
        lot = conn.execute(
            """SELECT l.*,p.code,p.name,p.unit FROM mu_lots l
               JOIN mu_products p ON p.id=l.product_id WHERE l.id=?""",
            (lot_id,),
        ).fetchone()
        if not lot:
            raise ValueError("ไม่พบ Lot")
        referenced = conn.execute(
            "SELECT 1 FROM mu_transaction_lots WHERE lot_id=? LIMIT 1", (lot_id,)
        ).fetchone()
        if referenced:
            raise ValueError("Lot นี้มีประวัติการเบิกแล้ว จึงลบไม่ได้ แต่สามารถแก้ไขข้อมูลพร้อมระบุเหตุผลได้")
        before = _available_stock(conn, lot["product_id"])
        after = before - int(lot["qty"])
        stamp = _now()
        actor = _current_admin_id()
        if int(lot["qty"]) > 0:
            conn.execute(
                """INSERT INTO mu_stock_events
                   (product_id,lot_id,event_type,qty,stock_before,stock_after,reason,actor_id,created_at)
                   VALUES(?,NULL,'delete_lot',?,?,?,?,?,?)""",
                (lot["product_id"], int(lot["qty"]), before, after, f"{reason}; Lot {lot['lot_number']}", actor, stamp),
            )
        conn.execute("DELETE FROM mu_lots WHERE id=?", (lot_id,))
        conn.execute(
            """INSERT INTO mu_audit_logs
               (actor_type,actor_id,action,entity_type,entity_id,detail,created_at)
               VALUES('admin',?,'delete_lot','lot',?,?,?)""",
            (actor, lot_id, f"ลบ {lot['code']} Lot {lot['lot_number']} จำนวน {lot['qty']}; เหตุผล: {reason}", stamp),
        )
        conn.commit()
        return _admin_response(True, "ลบ Lot และบันทึก Audit แล้ว")
    except (ValueError, sqlite3.Error) as exc:
        conn.rollback()
        return _admin_response(False, str(exc), status=400)
    finally:
        conn.close()


@mu.route("/admin/product/<int:product_id>/write-off", methods=["POST"])
@_admin_required
def write_off_product(product_id):
    _require_csrf()
    qty = request.form.get("qty", type=int)
    reason = _clean(request.form.get("reason"), 300)
    if not qty or qty <= 0 or not reason:
        return _admin_response(False, "กรุณาระบุจำนวนและเหตุผลการตัดจำหน่าย", status=400)
    conn = _connect()
    try:
        conn.execute("BEGIN IMMEDIATE")
        product = conn.execute(
            "SELECT id,code,name,unit,is_active FROM mu_products WHERE id=?",
            (product_id,),
        ).fetchone()
        if not product or not product["is_active"]:
            raise ValueError("ไม่พบสินค้าหรือสินค้าถูกปิดใช้งาน")
        before = _available_stock(conn, product_id)
        if qty > before:
            raise ValueError("จำนวนคงเหลือไม่พอสำหรับการตัดจำหน่าย")
        allocations = _consume_fifo(conn, product_id, qty)
        stamp = _now()
        actor = _current_admin_id()
        lot_names = []
        for lot_id, taken in allocations:
            lot = conn.execute("SELECT lot_number FROM mu_lots WHERE id=?", (lot_id,)).fetchone()
            lot_names.append(f"{lot['lot_number']} ({taken})")
            conn.execute(
                """INSERT INTO mu_stock_events
                   (product_id,lot_id,event_type,qty,stock_before,stock_after,reason,actor_id,created_at)
                   VALUES(?,?,'write_off',?,?,?,?,?,?)""",
                (product_id, lot_id, taken, before, before - qty, reason, actor, stamp),
            )
        conn.execute(
            """INSERT INTO mu_audit_logs
               (actor_type,actor_id,action,entity_type,entity_id,detail,created_at)
               VALUES('admin',?,'write_off','product',?,?,?)""",
            (actor, product_id, f"ตัดจำหน่าย {product['code']} {qty} {product['unit']}; Lot: {', '.join(lot_names)}; เหตุผล: {reason}", stamp),
        )
        conn.commit()
        return _admin_response(True, "ตัดจำหน่ายและหัก Lot แบบ FIFO แล้ว")
    except (ValueError, sqlite3.Error) as exc:
        conn.rollback()
        return _admin_response(False, str(exc), status=400)
    finally:
        conn.close()


@mu.route("/admin/transaction/<int:transaction_id>/edit", methods=["POST"])
@_admin_required
def edit_transaction(transaction_id):
    _require_csrf()
    new_qty = request.form.get("qty", type=int)
    reason = _clean(request.form.get("reason"), 300)
    if not new_qty or new_qty <= 0 or not reason:
        return _admin_response(False, "กรุณาระบุยอดใหม่และเหตุผล", tab="history", status=400)
    conn = _connect()
    try:
        conn.execute("BEGIN IMMEDIATE")
        tx = conn.execute("SELECT * FROM mu_transactions WHERE id=?", (transaction_id,)).fetchone()
        if not tx or tx["status"] != "Completed":
            raise ValueError("รายการนี้ไม่สามารถแก้ไขได้")
        old_qty = int(tx["qty"])
        _restore_allocations(conn, transaction_id)
        allocations = _consume_fifo(conn, tx["product_id"], new_qty)
        _save_allocations(conn, transaction_id, allocations)
        stamp = _now()
        actor = _current_admin_id()
        conn.execute(
            "UPDATE mu_transactions SET qty=?,updated_at=?,edited_by=? WHERE id=?",
            (new_qty, stamp, actor, transaction_id),
        )
        conn.execute(
            """INSERT INTO mu_audit_logs(actor_type,actor_id,action,entity_type,entity_id,detail,created_at)
               VALUES('admin',?,'edit','transaction',?,?,?)""",
            (actor, transaction_id, f"แก้ยอด {old_qty} -> {new_qty}; เหตุผล: {reason}", stamp),
        )
        conn.commit()
        return _admin_response(True, "แก้ไขยอดและคำนวณ FIFO ใหม่แล้ว", tab="history")
    except (ValueError, sqlite3.Error) as exc:
        conn.rollback()
        return _admin_response(False, str(exc), tab="history", status=400)
    finally:
        conn.close()


@mu.route("/admin/transaction/<int:transaction_id>/cancel", methods=["POST"])
@_admin_required
def cancel_transaction(transaction_id):
    _require_csrf()
    reason = _clean(request.form.get("reason"), 300)
    if not reason:
        return _admin_response(False, "กรุณาระบุเหตุผลการยกเลิก", tab="history", status=400)
    conn = _connect()
    try:
        conn.execute("BEGIN IMMEDIATE")
        tx = conn.execute("SELECT * FROM mu_transactions WHERE id=?", (transaction_id,)).fetchone()
        if not tx or tx["status"] != "Completed":
            raise ValueError("รายการนี้ถูกยกเลิกหรือไม่สามารถแก้ไขได้")
        _restore_allocations(conn, transaction_id)
        stamp = _now()
        actor = _current_admin_id()
        conn.execute(
            """UPDATE mu_transactions SET status='Cancelled',updated_at=?,cancelled_by=? WHERE id=?""",
            (stamp, actor, transaction_id),
        )
        conn.execute(
            """INSERT INTO mu_audit_logs(actor_type,actor_id,action,entity_type,entity_id,detail,created_at)
               VALUES('admin',?,'cancel','transaction',?,?,?)""",
            (actor, transaction_id, f"ยกเลิกรายการและคืน Lot; เหตุผล: {reason}", stamp),
        )
        conn.commit()
        return _admin_response(True, "ยกเลิกรายการและคืนสต็อกเข้า Lot เดิมแล้ว", tab="history")
    except (ValueError, sqlite3.Error) as exc:
        conn.rollback()
        return _admin_response(False, str(exc), tab="history", status=400)
    finally:
        conn.close()


@mu.route("/admin/employee/<emp_id>/password", methods=["POST"])
@_admin_required
def change_employee_password(emp_id):
    _require_csrf()
    password = request.form.get("password", "")
    if len(password) < 4:
        flash("รหัสผ่านหน้าเบิกต้องมีอย่างน้อย 4 ตัวอักษร", "danger")
        return redirect(url_for("mu.admin_dashboard", tab="employees"))
    conn = _connect()
    employee = conn.execute(
        """SELECT emp_id,name FROM users
           WHERE emp_id=? AND lower(trim(COALESCE(department,'')))='manufacturing'""",
        (emp_id,),
    ).fetchone()
    if not employee:
        conn.close()
        flash("ไม่พบพนักงาน Manufacturing ที่ต้องการจัดการ", "danger")
        return redirect(url_for("mu.admin_dashboard", tab="employees"))
    stamp = _now()
    actor = _current_admin_id()
    conn.execute(
        """INSERT INTO mu_user_access(emp_id,password_hash,updated_by,updated_at)
           VALUES(?,?,?,?)
           ON CONFLICT(emp_id) DO UPDATE SET
             password_hash=excluded.password_hash,
             updated_by=excluded.updated_by,
             updated_at=excluded.updated_at""",
        (emp_id, generate_password_hash(password), actor, stamp),
    )
    conn.execute(
        """INSERT INTO mu_audit_logs(actor_type,actor_id,action,entity_type,detail,created_at)
           VALUES('admin',?,'update_password','employee',?,?)""",
        (actor, f"ตั้งรหัสผ่านหน้าเบิกให้ {employee['name']} ({emp_id})", stamp),
    )
    conn.commit()
    conn.close()
    flash(f"ตั้งรหัสผ่านหน้าเบิกให้ {employee['name']} แล้ว", "success")
    return redirect(url_for("mu.admin_dashboard", tab="employees"))


@mu.route("/admin/profile", methods=["POST"])
@_admin_required
def update_admin_profile():
    _require_csrf()
    display_name = _clean(request.form.get("display_name"), 100)
    current_password = request.form.get("current_password", "")
    new_password = request.form.get("new_password", "")
    if not display_name or not current_password:
        flash("กรุณาระบุชื่อที่แสดงและรหัสผ่านปัจจุบัน", "danger")
        return redirect(url_for("mu.admin_dashboard", tab="profile"))
    if new_password and not 8 <= len(new_password) <= 128:
        flash("รหัสผ่านใหม่ต้องมี 8–128 ตัวอักษร", "danger")
        return redirect(url_for("mu.admin_dashboard", tab="profile"))
    conn = _connect()
    try:
        admin = conn.execute(
            "SELECT id,username,password,name,role FROM admins WHERE username=?",
            (_current_admin_id(),),
        ).fetchone()
        if (
            not admin
            or admin["role"] not in {"admin_mu", "superadmin"}
            or not check_password_hash(admin["password"], current_password)
        ):
            raise ValueError("รหัสผ่านปัจจุบันไม่ถูกต้อง")
        stamp = _now()
        if new_password:
            conn.execute(
                "UPDATE admins SET name=?,password=? WHERE id=?",
                (display_name, generate_password_hash(new_password), admin["id"]),
            )
        else:
            conn.execute("UPDATE admins SET name=? WHERE id=?", (display_name, admin["id"]))
        conn.execute(
            """INSERT INTO mu_audit_logs
               (actor_type,actor_id,action,entity_type,entity_id,detail,created_at)
               VALUES('admin',?,'update_profile','admin',?,?,?)""",
            (admin["username"], admin["id"], f"แก้ไขโปรไฟล์ {admin['name']} -> {display_name}", stamp),
        )
        conn.commit()
        session["admin_name"] = display_name
        flash("บันทึกโปรไฟล์แล้ว", "success")
    except ValueError as exc:
        conn.rollback()
        flash(str(exc), "danger")
    finally:
        conn.close()
    return redirect(url_for("mu.admin_dashboard", tab="profile"))


@mu.route("/admin/accounts", methods=["POST"])
@_admin_required
def add_mu_admin():
    _require_csrf()
    username = _clean(request.form.get("username"), 50).lower()
    display_name = _clean(request.form.get("display_name"), 100)
    password = request.form.get("password", "")
    if not username or not display_name or not 8 <= len(password) <= 128:
        flash("ข้อมูลบัญชีไม่ครบ หรือรหัสผ่านสั้นกว่า 8 ตัวอักษร", "danger")
        return redirect(url_for("mu.admin_dashboard", tab="admins"))
    if not all(char.isalnum() or char in "._-" for char in username):
        flash("Username ใช้ได้เฉพาะตัวอักษร ตัวเลข จุด ขีดกลาง และขีดล่าง", "danger")
        return redirect(url_for("mu.admin_dashboard", tab="admins"))
    conn = _connect()
    try:
        stamp = _now()
        cursor = conn.execute(
            "INSERT INTO admins(username,password,name,role) VALUES(?,?,?,'admin_mu')",
            (username, generate_password_hash(password), display_name),
        )
        conn.execute(
            """INSERT INTO mu_audit_logs
               (actor_type,actor_id,action,entity_type,entity_id,detail,created_at)
               VALUES('admin',?,'create_admin','admin',?,?,?)""",
            (_current_admin_id(), cursor.lastrowid, f"เพิ่มสิทธิ์ Admin MU ให้ {display_name} ({username})", stamp),
        )
        conn.commit()
        flash("เพิ่มบัญชี Admin MU แล้ว", "success")
    except sqlite3.IntegrityError:
        conn.rollback()
        flash("Username นี้ถูกใช้งานแล้ว", "danger")
    finally:
        conn.close()
    return redirect(url_for("mu.admin_dashboard", tab="admins"))


@mu.route("/admin/accounts/<int:admin_id>/delete", methods=["POST"])
@_admin_required
def delete_mu_admin(admin_id):
    _require_csrf()
    conn = _connect()
    try:
        target = conn.execute(
            "SELECT id,username,name,role FROM admins WHERE id=?", (admin_id,)
        ).fetchone()
        if not target or target["role"] != "admin_mu":
            raise ValueError("ไม่พบบัญชี Admin MU")
        if target["username"].casefold() == _current_admin_id().casefold():
            raise ValueError("ไม่สามารถลบบัญชีที่กำลังใช้งานอยู่")
        remaining = conn.execute(
            "SELECT COUNT(*) FROM admins WHERE role='admin_mu' AND id<>?", (admin_id,)
        ).fetchone()[0]
        if remaining < 1:
            raise ValueError("ระบบต้องเหลือ Admin MU อย่างน้อย 1 บัญชี")
        stamp = _now()
        conn.execute("DELETE FROM admins WHERE id=? AND role='admin_mu'", (admin_id,))
        conn.execute(
            """INSERT INTO mu_audit_logs
               (actor_type,actor_id,action,entity_type,entity_id,detail,created_at)
               VALUES('admin',?,'delete_admin','admin',?,?,?)""",
            (_current_admin_id(), admin_id, f"ลบสิทธิ์ Admin MU ของ {target['name']} ({target['username']})", stamp),
        )
        conn.commit()
        flash("ลบบัญชี Admin MU แล้ว", "success")
    except ValueError as exc:
        conn.rollback()
        flash(str(exc), "danger")
    finally:
        conn.close()
    return redirect(url_for("mu.admin_dashboard", tab="admins"))


@mu.route("/admin/product/<int:product_id>/lots")
@_admin_required
def product_lots(product_id):
    conn = _connect()
    rows = conn.execute(
        """SELECT id,lot_number,qty,received_date,expiry_date,created_by
           FROM mu_lots WHERE product_id=?
           ORDER BY date(received_date),id""", (product_id,)
    ).fetchall()
    conn.close()
    return jsonify([dict(row) for row in rows])
