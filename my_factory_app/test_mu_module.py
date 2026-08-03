import os
import sqlite3
import tempfile
import unittest
from datetime import datetime
from io import BytesIO

from flask import Flask, session
from openpyxl import load_workbook
from werkzeug.security import generate_password_hash

import mu_module


class MuModuleTestCase(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.db_path = os.path.join(self.temp_dir.name, "mu_test.db")
        os.environ["MU_ADMIN_USERNAME"] = "mu_test_admin"
        os.environ["MU_ADMIN_PASSWORD"] = "StrongPassword123!"
        os.environ["MU_USER_PASSWORD"] = "MuUserPass123!"

        app = Flask(__name__, template_folder="templates", static_folder="static")
        app.secret_key = "test-secret"
        app.jinja_env.globals["csrf_token"] = lambda: "test-csrf"
        app.add_url_rule("/", "index", lambda: "index")
        app.add_url_rule("/user-services", "user_services", lambda: "services")
        mu_module.register_mu_module(app, self.db_path, lambda: True)
        self.app = app
        self.client = app.test_client()

        conn = sqlite3.connect(self.db_path)
        conn.execute(
            """CREATE TABLE admins (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE, password TEXT, name TEXT, role TEXT
            )"""
        )
        conn.execute(
            "INSERT INTO admins(username,password,name,role) VALUES(?,?,?,'admin_mu')",
            ("mu_test_admin", generate_password_hash("StrongPassword123!"), "MU Test Admin"),
        )
        conn.execute(
            "INSERT INTO admins(username,password,name,role) VALUES(?,?,?,'superadmin')",
            ("super_test", generate_password_hash("SuperPassword123!"), "Super Test Admin"),
        )
        conn.execute(
            """CREATE TABLE users (
                emp_id TEXT PRIMARY KEY, name TEXT, department TEXT, location TEXT
            )"""
        )
        conn.execute(
            "INSERT INTO users VALUES ('E001','MU Tester','MU','PC1')"
        )
        conn.execute(
            "UPDATE users SET department='Manufacturing' WHERE emp_id='E001'"
        )
        conn.execute(
            "INSERT INTO users VALUES ('E002','Office Tester','Accounting','PC1')"
        )
        product_id = conn.execute(
            "SELECT id FROM mu_products WHERE code='MU-001'"
        ).fetchone()[0]
        conn.execute(
            """INSERT INTO mu_lots
               (product_id,lot_number,qty,received_date,created_by,created_at,updated_at)
               VALUES(?, 'OLD', 5, '2026-01-01', 'test', '2026-01-01', '2026-01-01')""",
            (product_id,),
        )
        conn.execute(
            """INSERT INTO mu_lots
               (product_id,lot_number,qty,received_date,created_by,created_at,updated_at)
               VALUES(?, 'NEW', 10, '2026-02-01', 'test', '2026-02-01', '2026-02-01')""",
            (product_id,),
        )
        conn.commit()
        conn.close()

    def tearDown(self):
        self.temp_dir.cleanup()
        for key in ("MU_ADMIN_USERNAME", "MU_ADMIN_PASSWORD", "MU_USER_PASSWORD"):
            os.environ.pop(key, None)

    def _login_user(self):
        with self.client.session_transaction() as user_session:
            user_session["user_id"] = "E001"
            user_session["mu_user_emp_id"] = "E001"

    def _login_admin(self):
        with self.client.session_transaction() as admin_session:
            admin_session["admin_logged_in"] = True
            admin_session["admin_username"] = "mu_test_admin"
            admin_session["admin_name"] = "MU Test Admin"
            admin_session["admin_role"] = "admin_mu"

    def _login_superadmin(self):
        with self.client.session_transaction() as admin_session:
            admin_session["admin_logged_in"] = True
            admin_session["admin_username"] = "super_test"
            admin_session["admin_name"] = "Super Test Admin"
            admin_session["admin_role"] = "superadmin"

    def test_superadmin_can_access_mu_admin_but_other_admin_roles_cannot(self):
        self._login_superadmin()
        response = self.client.get("/mu/admin")
        self.assertEqual(response.status_code, 200)

        with self.client.session_transaction() as admin_session:
            admin_session["admin_role"] = "admin_all"
        response = self.client.get("/mu/admin")
        self.assertEqual(response.status_code, 302)
        self.assertIn("admin=1", response.headers["Location"])

    def test_seed_contains_all_excel_items(self):
        conn = sqlite3.connect(self.db_path)
        self.assertEqual(conn.execute("SELECT COUNT(*) FROM mu_products").fetchone()[0], 140)
        conn.close()

    def test_only_manufacturing_user_can_access_and_admin_sets_password(self):
        with self.client.session_transaction() as user_session:
            user_session["user_id"] = "E002"
        response = self.client.get("/mu/access/E002")
        self.assertEqual(response.status_code, 302)

        with self.client.session_transaction() as user_session:
            user_session.clear()
            user_session["user_id"] = "E001"
        response = self.client.get("/mu/access/E001")
        self.assertEqual(response.status_code, 200)
        self.assertIn("ยังไม่มีสิทธิ์ใช้งาน MU".encode("utf-8"), response.data)
        self.assertNotIn('id="muPassword"'.encode(), response.data)

        self._login_admin()
        response = self.client.post(
            "/mu/admin/employee/E001/password",
            data={"csrf_token": "test-csrf", "password": "1234"},
        )
        self.assertEqual(response.status_code, 302)

        with self.client.session_transaction() as user_session:
            user_session.clear()
            user_session["user_id"] = "E001"
        response = self.client.post(
            "/mu/access/E001",
            data={"csrf_token": "test-csrf", "password": "1234"},
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("/mu/portal/E001", response.location)

    def test_direct_withdraw_uses_oldest_lot_first(self):
        self._login_user()
        conn = sqlite3.connect(self.db_path)
        product_id = conn.execute("SELECT id FROM mu_products WHERE code='MU-001'").fetchone()[0]
        conn.close()
        response = self.client.post(
            "/mu/withdraw/E001",
            data={"csrf_token": "test-csrf", "product_id": product_id, "qty": 7},
        )
        self.assertEqual(response.status_code, 302)
        conn = sqlite3.connect(self.db_path)
        lots = conn.execute(
            "SELECT lot_number,qty FROM mu_lots WHERE product_id=? ORDER BY received_date,id",
            (product_id,),
        ).fetchall()
        self.assertEqual(lots, [("OLD", 0), ("NEW", 8)])
        self.assertEqual(
            conn.execute("SELECT status FROM mu_transactions").fetchone()[0], "Completed"
        )
        conn.close()

    def test_batch_withdraw_is_atomic_and_uses_fifo(self):
        self._login_user()
        conn = sqlite3.connect(self.db_path)
        product_id = conn.execute(
            "SELECT id FROM mu_products WHERE code='MU-001'"
        ).fetchone()[0]
        conn.close()
        response = self.client.post(
            "/mu/withdraw-batch/E001",
            data={
                "csrf_token": "test-csrf",
                "items": f'[{{"productId":{product_id},"qty":6}}]',
            },
        )
        self.assertEqual(response.status_code, 302)
        conn = sqlite3.connect(self.db_path)
        lots = conn.execute(
            "SELECT lot_number,qty FROM mu_lots ORDER BY received_date,id"
        ).fetchall()
        self.assertEqual(lots, [("OLD", 0), ("NEW", 9)])
        conn.close()

    def test_cancel_restores_original_lots_and_records_actor(self):
        self.test_direct_withdraw_uses_oldest_lot_first()
        self._login_admin()
        response = self.client.post(
            "/mu/admin/transaction/1/cancel",
            data={"csrf_token": "test-csrf", "reason": "ทดสอบคืนสินค้า"},
        )
        self.assertEqual(response.status_code, 302)
        conn = sqlite3.connect(self.db_path)
        lots = conn.execute(
            "SELECT lot_number,qty FROM mu_lots ORDER BY received_date,id"
        ).fetchall()
        self.assertEqual(lots, [("OLD", 5), ("NEW", 10)])
        tx = conn.execute(
            "SELECT status,cancelled_by FROM mu_transactions WHERE id=1"
        ).fetchone()
        self.assertEqual(tx, ("Cancelled", "mu_test_admin"))
        self.assertEqual(
            conn.execute(
                "SELECT COUNT(*) FROM mu_audit_logs WHERE action='cancel' AND actor_id='mu_test_admin'"
            ).fetchone()[0],
            1,
        )
        conn.close()

    def test_admin_can_manage_product_lot_writeoff_and_audit(self):
        self._login_admin()
        conn = sqlite3.connect(self.db_path)
        product_id = conn.execute(
            "SELECT id FROM mu_products WHERE code='MU-001'"
        ).fetchone()[0]
        lot_id = conn.execute(
            "SELECT id FROM mu_lots WHERE product_id=? ORDER BY id LIMIT 1",
            (product_id,),
        ).fetchone()[0]
        conn.close()

        response = self.client.post(
            f"/mu/admin/product/{product_id}/edit",
            data={
                "csrf_token": "test-csrf",
                "code": "MU-001",
                "name": "Edited MU Item",
                "unit": "PCS",
                "safety_stock": 3,
            },
        )
        self.assertEqual(response.status_code, 302)
        response = self.client.post(
            f"/mu/admin/lot/{lot_id}/edit",
            data={
                "csrf_token": "test-csrf",
                "lot_number": "OLD-EDIT",
                "qty": 6,
                "received_date": "2026-01-01",
                "expiry_date": "",
                "reason": "ตรวจนับจริง",
            },
        )
        self.assertEqual(response.status_code, 302)
        response = self.client.post(
            f"/mu/admin/product/{product_id}/write-off",
            data={"csrf_token": "test-csrf", "qty": 2, "reason": "ชำรุด"},
        )
        self.assertEqual(response.status_code, 302)

        conn = sqlite3.connect(self.db_path)
        self.assertEqual(
            conn.execute("SELECT name FROM mu_products WHERE id=?", (product_id,)).fetchone()[0],
            "Edited MU Item",
        )
        self.assertEqual(
            conn.execute("SELECT COALESCE(SUM(qty),0) FROM mu_lots WHERE product_id=?", (product_id,)).fetchone()[0],
            14,
        )
        self.assertGreaterEqual(
            conn.execute("SELECT COUNT(*) FROM mu_stock_events WHERE product_id=?", (product_id,)).fetchone()[0],
            2,
        )
        conn.close()

    def test_admin_profile_and_mu_admin_account_scope(self):
        self._login_admin()
        response = self.client.post(
            "/mu/admin/profile",
            data={
                "csrf_token": "test-csrf",
                "display_name": "Updated MU Admin",
                "current_password": "StrongPassword123!",
                "new_password": "",
            },
        )
        self.assertEqual(response.status_code, 302)
        response = self.client.post(
            "/mu/admin/accounts",
            data={
                "csrf_token": "test-csrf",
                "username": "mu_second",
                "display_name": "Second MU Admin",
                "password": "SecondStrong123!",
            },
        )
        self.assertEqual(response.status_code, 302)
        conn = sqlite3.connect(self.db_path)
        row = conn.execute(
            "SELECT name,role FROM admins WHERE username='mu_second'"
        ).fetchone()
        self.assertEqual(row, ("Second MU Admin", "admin_mu"))
        conn.close()

    def test_all_admin_tabs_render(self):
        self._login_admin()
        for tab in ("stock", "history", "audit", "monthly", "employees", "admins", "profile"):
            with self.subTest(tab=tab):
                response = self.client.get(f"/mu/admin?tab={tab}")
                self.assertEqual(response.status_code, 200)

    def test_delete_lot_rejects_referenced_history(self):
        self.test_direct_withdraw_uses_oldest_lot_first()
        self._login_admin()
        conn = sqlite3.connect(self.db_path)
        referenced_lot_id = conn.execute(
            "SELECT lot_id FROM mu_transaction_lots ORDER BY lot_id LIMIT 1"
        ).fetchone()[0]
        conn.close()
        response = self.client.post(
            f"/mu/admin/lot/{referenced_lot_id}/delete",
            data={"csrf_token": "test-csrf", "reason": "ทดสอบลบ"},
        )
        self.assertEqual(response.status_code, 302)
        conn = sqlite3.connect(self.db_path)
        self.assertEqual(
            conn.execute("SELECT COUNT(*) FROM mu_lots WHERE id=?", (referenced_lot_id,)).fetchone()[0],
            1,
        )
        conn.close()

    def test_auto_product_code_auto_lot_and_ajax_response(self):
        self._login_admin()
        response = self.client.get("/mu/admin/next-product-code")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["next_code"], "MU-146")

        response = self.client.post(
            "/mu/admin/product",
            data={
                "csrf_token": "test-csrf",
                "code": "",
                "name": "Auto Number Item",
                "unit": "PCS",
                "safety_stock": 0,
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.get_json()["success"])
        conn = sqlite3.connect(self.db_path)
        product_id = conn.execute(
            "SELECT id FROM mu_products WHERE code='MU-146'"
        ).fetchone()[0]
        conn.close()

        response = self.client.get(
            f"/mu/admin/product/{product_id}/next-lot?received_date=2026-07-31"
        )
        self.assertEqual(response.get_json()["next_lot"], "MU-146-20260731-001")
        response = self.client.post(
            "/mu/admin/lot",
            data={
                "csrf_token": "test-csrf",
                "product_id": product_id,
                "lot_number": "",
                "qty": 4,
                "received_date": "2026-07-31",
                "expiry_date": "",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.get_json()["success"])
        conn = sqlite3.connect(self.db_path)
        self.assertEqual(
            conn.execute(
                "SELECT lot_number FROM mu_lots WHERE product_id=?", (product_id,)
            ).fetchone()[0],
            "MU-146-20260731-001",
        )
        conn.close()

    def test_audit_excel_export_has_expected_sheets(self):
        self._login_admin()
        response = self.client.get("/mu/admin/audit/export.xlsx")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data.startswith(b"PK"))
        workbook = load_workbook(BytesIO(response.data), read_only=True)
        monthly_sheet_name = f"รายเดือน {datetime.now().strftime('%Y-%m')}"
        self.assertEqual(
            workbook.sheetnames,
            ["รายการเบิก", "Stock Movement", "Audit Log", monthly_sheet_name],
        )
        self.assertEqual(workbook["รายการเบิก"]["A1"].value, "วันเวลา")
        monthly_sheet = workbook[monthly_sheet_name]
        self.assertEqual(
            [monthly_sheet.cell(1, column).value for column in range(1, 10)],
            ["No.", "รหัสสินค้า", "รายการ / List", "รวม", "ยอดเบิก", "คงเหลือ", "Safety Stock", "สั่งซื้อ", "หน่วย"],
        )
        self.assertNotIn("ยอดยกมา", [cell.value for cell in monthly_sheet[1]])
        self.assertNotIn("รับเข้า", [cell.value for cell in monthly_sheet[1]])
        self.assertNotIn("ปรับยอด", [cell.value for cell in monthly_sheet[1]])
        workbook.close()

    def test_add_product_can_create_atomic_initial_lot(self):
        self._login_admin()
        response = self.client.post(
            "/mu/admin/product",
            data={
                "csrf_token": "test-csrf",
                "code": "",
                "name": "Initial Stock Item",
                "unit": "PCS",
                "safety_stock": 3,
                "initial_qty": 10,
                "received_date": "2026-07-31",
                "expiry_date": "2027-07-31",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.get_json()["success"])
        conn = sqlite3.connect(self.db_path)
        row = conn.execute(
            """SELECT p.code,l.lot_number,l.qty,l.received_date,l.expiry_date
               FROM mu_products p JOIN mu_lots l ON l.product_id=p.id
               WHERE p.code='MU-146'"""
        ).fetchone()
        conn.close()
        self.assertEqual(
            row,
            ("MU-146", "MU-146-20260731-001", 10, "2026-07-31", "2027-07-31"),
        )

    def test_auto_lot_sequence_is_compatible_with_fifo_tie_break(self):
        self._login_admin()
        conn = sqlite3.connect(self.db_path)
        product_id = conn.execute(
            "SELECT id FROM mu_products WHERE code='MU-002'"
        ).fetchone()[0]
        conn.close()
        for qty in (3, 4):
            response = self.client.post(
                "/mu/admin/lot",
                data={
                    "csrf_token": "test-csrf",
                    "product_id": product_id,
                    "lot_number": "",
                    "qty": qty,
                    "received_date": "2026-07-31",
                    "expiry_date": "",
                },
                headers={"X-Requested-With": "XMLHttpRequest"},
            )
            self.assertEqual(response.status_code, 200)

        conn = sqlite3.connect(self.db_path)
        lots = conn.execute(
            "SELECT lot_number,qty FROM mu_lots WHERE product_id=? ORDER BY received_date,id",
            (product_id,),
        ).fetchall()
        self.assertEqual(
            [row[0] for row in lots],
            ["MU-002-20260731-001", "MU-002-20260731-002"],
        )
        conn.close()

        self._login_user()
        response = self.client.post(
            "/mu/withdraw/E001",
            data={"csrf_token": "test-csrf", "product_id": product_id, "qty": 5},
        )
        self.assertEqual(response.status_code, 302)
        conn = sqlite3.connect(self.db_path)
        remaining = conn.execute(
            "SELECT lot_number,qty FROM mu_lots WHERE product_id=? ORDER BY received_date,id",
            (product_id,),
        ).fetchall()
        self.assertEqual(
            remaining,
            [("MU-002-20260731-001", 0), ("MU-002-20260731-002", 2)],
        )
        conn.close()


if __name__ == "__main__":
    unittest.main()
